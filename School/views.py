
import os
from django.http import HttpResponse
from django.conf import settings
import pandas as pd
import random
from openpyxl import load_workbook
from django.http import HttpResponseBadRequest
import openai
import logging

# Configuration du logger

logger = logging.getLogger('app')

#############################################################
#############################################################
#############################################################
#######                                             #########
#######                                             #########
#######                 SCHOOL DASHBOARD            #########
#######                                             #########
#######                                             #########
#############################################################
#############################################################
#############################################################
#SCHOOL_DASHBOARD
def home_school_dashboard(request):
   return render(request, 'dashboard/school/home.html')

def school_dashboard(request):
    context = {
        'username': request.user.username,
        'email': request.user.email,
        'school': request.user.school_name,
        'photo': request.user.photo,
        'role': 'School Leader/Principal'  # Replace with actual role if available
    }
    return render(request, 'dashboard/school/school_dashboard.html', context)

def teachers_school_dashboard(request):

    return render(request, 'dashboard/school/teachers.html')

def parents_school_dashboard(request):
    return render(request, 'dashboard/school/parents.html')

def students_school_dashboard(request):
    return render(request, 'dashboard/school/students.html')

def communicate_school_dashboard(request):
    return render(request, 'dashboard/school/communicate.html')



#############################################################
#############################################################
#############################################################
#######                                             #########
#######                                             #########
#######                 STUDENT DASHBOARD           #########
#######                                             #########
#######                                             #########
#############################################################
#############################################################
#############################################################
#STUDENT_DASHBOARD
def class_student_dashboard(request):
   return render(request, 'dashboard/student/class.html')

def student_dashboard(request):
    context = {
        'username': request.user.username,
        'firstname': request.user.first_name,
        'lastname': request.user.last_name,
        'email': request.user.email,
        'photo': request.user.photo,
    }
    return render(request, 'dashboard/student/student_dashboard.html', context)

def assignments_student_dashboard(request):
   return render(request, 'dashboard/student/assignments.html')

def gradebook_student_dashboard(request):
   return render(request, 'dashboard/student/gradebook.html')

def attendance_student_dashboard(request):
   return render(request, 'dashboard/student/attendance.html')

def character_student_dashboard(request):
   return render(request, 'dashboard/student/character.html')

def communicate_student_dashboard(request):
    return render(request, 'dashboard/student/communicate.html')

def library_student_dashboard(request):
    return render(request, 'dashboard/student/library.html')

def reward_student_dashboard(request):
    return render(request, 'dashboard/student/reward.html')

def tutor_student_dashboard(request):
    return render(request, 'dashboard/student/tutor.html')


#############################################################
#############################################################
#############################################################
#######                                             #########
#######                                             #########
#######                 PARENT DASHBOARD            #########
#######                                             #########
#######                                             #########
#############################################################
#############################################################
#############################################################
#PARENT_DASHBOARD
def class_parent_dashboard(request):
   return render(request, 'dashboard/parent/class.html')

def parent_dashboard(request):
    context = {
        'username': request.user.username,
        'firstname': request.user.first_name,
        'lastname': request.user.last_name,
        'email': request.user.email,
        'photo': request.user.photo,
    }
    return render(request, 'dashboard/parent/parent_dashboard.html', context)

def assignments_parent_dashboard(request):
   return render(request, 'dashboard/parent/assignments.html')

def gradebook_parent_dashboard(request):
   return render(request, 'dashboard/parent/gradebook.html')

def attendance_parent_dashboard(request):
   return render(request, 'dashboard/parent/attendance.html')

def communicate_parent_dashboard(request):
    return render(request, 'dashboard/parent/communicate.html')

def library_parent_dashboard(request):
    return render(request, 'dashboard/parent/library.html')


#############################################################
#############################################################
#############################################################
#######                                             #########
#######                                             #########
#######                 TEAHCER DASHBOARD           #########
#######                                             #########
#######                                             #########
#############################################################
#############################################################
#############################################################

#############################################################
#######                                             #########
#######                                             #########
#######                  NAVIGATION MENU            #########
#######                                             #########
#######                                             #########
#############################################################
# USER FLOW, UPDATE USER INFORMATIONS
def update_profile_photo(request):
    if request.method == "POST":
        user = request.user
        profile_photo = request.FILES.get('profile_picture')  # Récupérer le fichier sélectionné

        if profile_photo:
            user.photo = profile_photo
            user.save()  # Enregistrer la mise à jour
            messages.success(request, "Profile picture updated successfully!")
        else:
            messages.error(request, "Please select a valid image.")

        # Rediriger ou recharger la page en fonction du type d'utilisateur
        return redirect('teacher_dashboard')

    return render(request, 'dashboard/teacher/teacher_dashboard.html')

# DASHBOARD PAGE OF THE TEAHCER
def teacher_dashboard(request):

    if request.method == "POST":
        user = request.user
        profile_photo = request.FILES.get('profile_picture')  # Récupérer le fichier sélectionné

        if profile_photo:
            user.photo = profile_photo
            user.save()  # Enregistrer la mise à jour
            messages.success(request, "Profile picture updated successfully!")
            context = {
                'username': request.user.username,
                'firstname': request.user.first_name,
                'lastname': request.user.last_name,
                'email': request.user.email,
                'role_user': request.user.user_type,
                'school': request.user.school_name,
                'photo': request.user.photo,
            }
            # Rediriger ou recharger la page en fonction du type d'utilisateur
            return render(request, 'dashboard/teacher/teacher_dashboard.html', context)
        else:
            messages.error(request, "Please select a valid image.")


    context = {
        'username': request.user.username,
        'firstname': request.user.first_name,
        'lastname': request.user.last_name,
        'email': request.user.email,
        'role_user': request.user.user_type,
        'school': request.user.school_name,
        'photo': request.user.photo,
    }
    return render(request, 'dashboard/teacher/teacher_dashboard.html',context)

# ASSIGNMENTS PAGE OF THE TEAHCER
def assignments_teacher_dashboard(request):
   return render(request, 'dashboard/teacher/assignments.html')

# ATTENDANCE PAGE OF THE TEAHCER
def attendance_teacher_dashboard(request):
   return render(request, 'dashboard/teacher/attendance.html')

# CHARACTER PAGE OF THE TEAHCER
def character_teacher_dashboard(request):
   return render(request, 'dashboard/teacher/character.html')

# COMMUNICATE PAGE OF THE TEAHCER
def communicate_teacher_dashboard(request):
    return render(request, 'dashboard/teacher/communicate.html')

# GRADEBOOK PAGE OF THE TEAHCER
def gradebook_teacher_dashboard(request):
   return render(request, 'dashboard/teacher/gradebook.html')

# POLLS PAGE OF THE TEAHCER
def polis_teacher_dashboard(request):
    return render(request, 'dashboard/teacher/polis.html')

#############################################################
#######                                             #########
#######                                             #########
#######            AI_CHAT PAGE OF THE TEAHCER      #########
#######                                             #########
#######                                             #########
#############################################################

openai.api_key = settings.OPENAI_API_KEY  # Récupérer la clé API depuis les paramètres
def ai_chat_teacher_dashboard(request):
    return render(request, 'dashboard/teacher/ai_chat.html')

def chatbot_view(request):
    return render(request, "chatbot/chat.html")

def chat_api(request):
    if request.method == "POST":
        user_message = request.POST.get("message")
        user = request.user if request.user.is_authenticated else None

        # Mise à jour de l'appel API pour correspondre à la nouvelle interface
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": user_message}]
        )

        # Récupérer la réponse du bot
        bot_response = response.choices[0].message.content

        # Enregistrer l'historique de chat pour les utilisateurs authentifiés
        # if user:
        # ChatHistory.objects.create(user=user, message=user_message, response=bot_response)

        return JsonResponse({"response": bot_response})

# REPORTS PAGE OF THE TEACHER
def reports_teacher_dashboard(request):
    return render(request, 'dashboard/teacher/reports.html')

# SEATING PAGE OF THE TEAHCER
def seating_teacher_dashboard(request):
    return render(request, 'dashboard/teacher/seating.html')

# TEACH PAGE OF THE TEAHCER
def teach_teacher_dashboard(request):
    return render(request, 'dashboard/teacher/teach.html')



#############################################################
#######                                             #########
#######                                             #########
#######    DIGITAL LIBRARY PAGE OF THE TEACHER       #########
#######                                             #########
#######                                             #########
#############################################################

import fitz
import requests
from io import BytesIO
import qrcode
from django.db import connection
import zipfile
from zipfile import ZipFile
import shutil
from datetime import date
from django.contrib import messages

# Function to filter the questions from the EXCEL SHEET OF MATHEMATICS
def filter_math_question(subject, number, grade, domain, kind):

    df = pd.read_excel("Digital library2.xlsx")

    filtered_df = df.loc[
        (df["Subject"] == subject) &
        (df["Grade"] == grade) &
        (df["Domain"] == domain) &
        (df["MC/OR"] == kind)
    ]

    workbook = load_workbook("Digital library2.xlsx")
    sheet = workbook.active

    links = []

    column = df.columns.get_loc("Link to item")
    try:
        q = random.sample(list(filtered_df.index), number)
    except ValueError:
        raise ValueError("Not enough questions to generate test")

    for ele in q:
        cell = sheet.cell(ele+2, column+1)
        link = cell.hyperlink.target
        links.append(link)
    return links

# Function to filter the questions from the EXCEL SHEET OF LANGUAGES
def filter_lang_question(subject, number, grade, domain,  kind):
    df = pd.read_excel("Dreametrix excel.xlsx")

    filtered_df = df.loc[
        (df["Subject"] == subject) &
        (df["Grade"] == grade) &
        (df["Domain"] == domain) &
        (df["MC/OR"] == kind)
    ]

    workbook = load_workbook("Dreametrix excel.xlsx")
    sheet = workbook.active


    story_column = df.columns.get_loc("Story")
    stories = random.sample(list(filtered_df.index), number)


    story_links = {}

    def _get_question(story):
        res = []
        value = filtered_df.loc[story, "Story"]
        rows = filtered_df.index[filtered_df["Link to item"] == value].tolist()
        for row in rows:
            cell = sheet.cell(row+2, story_column+2)
            link = cell.hyperlink.target
            res.append(link)
        return res

    for story in stories:
        q = tuple(_get_question(story))

        cell = sheet.cell(story+2, story_column+1)
        link = cell.hyperlink.target
        story_links[link] = q

    return story_links

def generate_pdf(links, selected_class, subject, grade, teacher_name, student_id, assignment_type, domain, generate_answer_sheet_flag=False):
    """ Génère un PDF pour un élève spécifique """
    logger.info('Génération du PDF pour un élève spécifique')

    student = get_object_or_404(Student, id=student_id)
    student_name = f"{student.user.first_name} {student.user.last_name}"

    doc = fitz.open()

    # === Page de couverture ===
    cover = doc.new_page(width=595, height=842)

    # Fonction pour centrer le texte
    def center_text(page, text, y, fontsize, fontname="helv", color=(0, 0, 0)):
        text_width = fitz.get_text_length(text, fontname=fontname, fontsize=fontsize)
        x = (page.rect.width - text_width) / 2  # Centrer horizontalement
        page.insert_text((x, y), text, fontsize=fontsize, fontname=fontname, color=color)

    # Ajouter les éléments centrés
    center_text(cover, f"GRADE {grade} {subject.upper()}", 200, fontsize=20)
    center_text(cover, f"Mr. {teacher_name}", 250, fontsize=14)
    center_text(cover, f"{assignment_type} : {domain}", 300, fontsize=16)
    cover.insert_text((100, 650), f"Name: {student_name}", fontsize=12)
    cover.insert_text((100, 680), f"Class: {selected_class}", fontsize=12)

    # === Pages de questions ===
    question_number = 1
    if isinstance(links, list):
        for i, link in enumerate(links):
            link = link.replace("dl=0", "raw=1")

            # Créer une nouvelle page après chaque deuxième question
            if i % 2 == 0:
                current_page = doc.new_page(width=595, height=842)
                y_top = 50  # Position Y pour la première question
                y_bottom = 421  # Position Y pour la deuxième question (50% de la hauteur)

            try:
                response = requests.get(link)
                image = BytesIO(response.content)

                # === Définir l'espace pour la question ===
                if i % 2 == 0:
                    y = y_top  # Première question en haut
                else:
                    y = y_bottom  # Deuxième question en bas

                # === Ajout du fond gris pour le numéro ===
                num_size = 18
                rect_size = 25  # Taille du carré de fond
                rect = fitz.Rect(50, y, 50 + rect_size, y + rect_size)
                current_page.draw_rect(rect, color=(0.8, 0.8, 0.8), fill=(0.8, 0.8, 0.8))

                # === Insérer le numéro de la question au centre du carré ===
                question_text = f"{question_number}"
                text_x = 50 + 7  # Ajustement pour centrer le texte
                text_y = y + 17
                current_page.insert_text((text_x, text_y), question_text, fontsize=num_size, color=(0, 0, 0), fontname="Helvetica-Bold")

                # === Définir l'espace pour l'image ===
                img_x = 50 + rect_size + 10  # Espace entre le numéro et l'image (aligné à gauche)
                img_y = y - 3  # Alignement vertical avec le numéro
                img_width = 480  # Largeur fixe pour l'image
                img_height = 340  # Hauteur fixe pour l'image (50% de la hauteur de la page moins les marges)

                # === Insérer l’image (alignée à gauche) ===
                rect = fitz.Rect(img_x, img_y, img_x + img_width, img_y + img_height)
                current_page.insert_image(rect, stream=image)

                question_number += 1

            except Exception as e:
                raise ValueError(e)

        # === Génération du QR Code personnalisé ===
        try:
            class_instance = Class.objects.get(name=selected_class)
            qr_data = f"student_id:{student_id}|class_id:{class_instance.id}|assignment_type:{assignment_type}"

            # Création du QR Code de base
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=6, border=2)
            qr.add_data(qr_data)
            qr.make(fit=True)

            # Création de l'image QR code avec couleur personnalisée
            qr_img = qr.make_image(fill_color=(62, 179, 227), back_color="white").convert("RGBA")

            # Convertir l'image en BytesIO pour une insertion dans un PDF
            qr_bytes = BytesIO()
            qr_img.save(qr_bytes, format='PNG')
            qr_bytes.seek(0)

        except Exception as e:
            logger.error(f"Erreur lors de la création du QR code stylisé : {e}")

        # === Ajouter le QR Code à chaque page ===
        for page in doc:
            page.insert_image(fitz.Rect(521, 768, 571, 818), stream=qr_bytes)
            qr_bytes.seek(0)

        # === Ajouter la feuille de réponses si la case est cochée ===
            # === Ajouter la feuille de réponses si la case est cochée ===
            if generate_answer_sheet_flag:
                answer_sheet_page = doc.new_page(width=595, height=842)
                answer_sheet_page.insert_text((50, 30), f"Name: {student_name}", fontsize=12, fontname="helv",
                                              color=(0, 0, 0))
                answer_sheet_page.insert_text((50, 50), f"Class: {selected_class}", fontsize=12, fontname="helv",
                                              color=(0, 0, 0))
                answer_sheet_page.insert_text((50, 70), f"Date: {date.today().strftime('%Y-%m-%d')}", fontsize=12,
                                              fontname="helv", color=(0, 0, 0))

                start_y = 110  # Position Y de départ pour les questions
                increment_y = 30  # Espacement vertical entre les questions
                num_questions = len(links)  # Nombre de questions dynamique
                questions_per_column = (num_questions + 1) // 2  # Nombre de questions par colonne

                # Colonne 1
                for i in range(questions_per_column):
                    question_number = i + 1
                    answer_sheet_page.insert_text((50, start_y + i * increment_y), f"{question_number}.", fontsize=12,
                                                  fontname="helv", color=(0, 0, 0))

                    # Ajouter les cases carrées avec les lettres A, B, C, D
                    square_size = 15  # Taille de chaque case
                    letters = ['A', 'B', 'C', 'D']
                    for j, letter in enumerate(letters):
                        x = 100 + j * (square_size + 10)  # Espacement horizontal entre les cases
                        y = start_y + i * increment_y - 5  # Ajustement vertical pour aligner avec le texte
                        rect = fitz.Rect(x, y, x + square_size, y + square_size)
                        answer_sheet_page.draw_rect(rect, color=(0, 0, 0), fill=None)  # Contour noir
                        answer_sheet_page.insert_text((x + 5, y + 10), letter, fontsize=10, fontname="helv",
                                                      color=(0, 0, 0))  # Lettre au centre

                # Colonne 2
                for i in range(questions_per_column, num_questions):
                    question_number = i + 1
                    answer_sheet_page.insert_text((300, start_y + (i - questions_per_column) * increment_y),
                                                  f"{question_number}.", fontsize=12, fontname="helv", color=(0, 0, 0))

                    # Ajouter les cases carrées avec les lettres A, B, C, D
                    square_size = 15  # Taille de chaque case
                    letters = ['A', 'B', 'C', 'D']
                    for j, letter in enumerate(letters):
                        x = 350 + j * (square_size + 10)  # Espacement horizontal entre les cases
                        y = start_y + (
                                    i - questions_per_column) * increment_y - 5  # Ajustement vertical pour aligner avec le texte
                        rect = fitz.Rect(x, y, x + square_size, y + square_size)
                        answer_sheet_page.draw_rect(rect, color=(0, 0, 0), fill=None)  # Contour noir
                        answer_sheet_page.insert_text((x + 5, y + 10), letter, fontsize=10, fontname="helv",
                                                      color=(0, 0, 0))  # Lettre au centre

                # Section pour les remarques
                answer_sheet_page.insert_text((50, start_y + questions_per_column * increment_y + 20),
                                              "Remarks: __________________________________________________",
                                              fontsize=12, fontname="helv", color=(0, 0, 0))

                # Ajouter le QR Code à la dernière page (answer sheet)
                answer_sheet_page.insert_image(fitz.Rect(521, 768, 571, 818), stream=qr_bytes)

            pdf_filename = f"test_{student_name}.pdf"
            logger.info(f"PDF généré : {pdf_filename}")
            doc.save(pdf_filename)
            return pdf_filename

# =============================== #
# Vue Django pour la génération   #
# =============================== #

def generate_pdf_view(request):
    logger.info("LOG Requête reçue pour la génération du PDF")

    if request.method == "POST":
        selected_class = request.POST.get('classes', '')
        if not selected_class:
            messages.error(request, "Please select a class.")
            return render(request, 'dashboard/teacher/digital_library.html')

        teacher_name = request.user.get_full_name() or request.user.username
        logger.info(f"Name of the professor : {teacher_name}")

        subject = request.POST['subject']
        number = int(request.POST['number'])
        grade = int(request.POST['grade'])
        kind = request.POST['kind']
        domain = request.POST['domain']
        assignment_type = request.POST.get('assignment_type', 'Quiz')
        generate_answer_sheet_flag = request.POST.get('generateAnswerSheet', False) == 'on'

        try:
            # Générer les liens de questions selon la matière
            if subject == "Math":
                links = filter_math_question(subject, number, grade, domain, kind)
            else:
                links = filter_lang_question(subject, number, grade, domain, kind)

            pdf_files = []
            temp_dir = f"quiz_class_{selected_class}"  # Dossier temporaire pour stocker les PDFs
            os.makedirs(temp_dir, exist_ok=True)

            # Générer les fichiers PDF pour chaque élève
            class_instance = Class.objects.get(name=selected_class)
            logger.info(f"Classe sélectionnée : {selected_class}")
            logger.info(f"Classe ID : {class_instance.id}")

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT sc."student_id"
                    FROM "public"."Authentication_user" AS u
                    INNER JOIN "public"."Authentication_student" AS s ON u."id" = s."user_id"
                    INNER JOIN "public"."School_class_students" AS sc ON s."id" = sc."student_id"
                    WHERE sc."class_id" = %s;
                """, [class_instance.id])

                results = cursor.fetchall()
                logger.info(f"Nombre d'étudiants récupérés : {len(results)}")

            # Générer les PDFs pour les quiz
            for student in results:
                student_id = student[0]
                logger.info(f"Début de génération du PDF pour l'élève ID : {student_id}")
                pdf_filename = generate_pdf(
                    links=links,
                    selected_class=selected_class,
                    subject=subject,
                    grade=grade,
                    teacher_name=teacher_name,
                    student_id=student_id,
                    domain=domain,
                    assignment_type=assignment_type,
                    generate_answer_sheet_flag=generate_answer_sheet_flag
                )
                if not os.path.exists(pdf_filename) or os.path.getsize(pdf_filename) == 0:
                    logger.error(f"Le fichier PDF {pdf_filename} est vide ou corrompu !")
                else:
                    logger.info(f"PDF généré avec succès : {pdf_filename}")
                    pdf_files.append(pdf_filename)
                    shutil.move(pdf_filename, os.path.join(temp_dir, os.path.basename(pdf_filename)))

            # Vérification : A-t-on au moins un fichier PDF valide ?
            if not pdf_files:
                messages.error(request, "Some issues occured : No valide PDF generated.")
                shutil.rmtree(temp_dir)  # Nettoyage
                return render(request, 'dashboard/teacher/digital_library.html')

            # Création du fichier ZIP
            zip_filename = os.path.join(temp_dir, f"quiz_class_{selected_class}.zip")
            with ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for pdf in pdf_files:
                    zipf.write(os.path.join(temp_dir, os.path.basename(pdf)), arcname=os.path.basename(pdf))

            # Vérification : Le ZIP a-t-il été bien créé ?
            if not os.path.exists(zip_filename) or os.path.getsize(zip_filename) == 0:
                logger.error("Le fichier ZIP n'a pas été correctement généré !")
                messages.error(request, "Erreur lors de la création du fichier ZIP.")
                shutil.rmtree(temp_dir)  # Nettoyage
                return render(request, 'dashboard/teacher/digital_library.html')

            # Envoi du fichier ZIP en réponse HTTP
            response = HttpResponse(open(zip_filename, "rb"), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(zip_filename)}"'

            # Nettoyage des fichiers temporaires après envoi
            shutil.rmtree(temp_dir)

            return response

        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'dashboard/teacher/digital_library.html')

    return render(request, 'dashboard/teacher/digital_library.html')

def generate_pdf_preview(request):
    if request.method == "POST":
        logger.info("Début de la génération du preview")
        try:
            # Extraire les données du formulaire
            selected_class = request.POST.get('classes', '')
            subject = request.POST['subject']
            number = int(request.POST['number'])
            grade = int(request.POST['grade'])
            kind = request.POST['kind']
            domain = request.POST['domain']
            teacher_name = request.user.get_full_name() or request.user.username
            generate_answer_sheet_flag = request.POST.get('generateAnswerSheet', False) == 'on'  # Récupérer la valeur du checkbox

            logger.info(f"Données du formulaire : subject={subject}, number={number}, grade={grade}, kind={kind}, domain={domain}")

            # Vérifier si une classe est sélectionnée
            if not selected_class:
                raise Exception("Aucune classe sélectionnée.")

            # Générer les liens de questions selon la matière
            if subject == "Math":
                links = filter_math_question(subject, number, grade, domain, kind)
            else:
                links = filter_lang_question(subject, number, grade, domain, kind)

            logger.info(f"Liens générés : {links}")

            # Récupérer un étudiant de la classe sélectionnée
            class_instance = Class.objects.get(name=selected_class)
            logger.info(f"Classe sélectionnée : {selected_class}")
            logger.info(f"Classe ID : {class_instance.id}")

            with connection.cursor() as cursor:
                # Exécution de la requête SQL brute pour récupérer un étudiant
                cursor.execute("""
                    SELECT sc."student_id"
                    FROM "public"."Authentication_user" AS u
                    INNER JOIN "public"."Authentication_student" AS s ON u."id" = s."user_id"
                    INNER JOIN "public"."School_class_students" AS sc ON s."id" = sc."student_id"
                    WHERE sc."class_id" = %s;
                """, [class_instance.id])

                # Récupérer le premier étudiant
                student = cursor.fetchone()

                if student:
                    student_id = student[0]
                    logger.info(f"Étudiant récupéré pour le preview : ID={student_id}")
                else:
                    # Aucun étudiant trouvé, utiliser un nom fictif
                    student_id = None
                    logger.warning("Aucun étudiant trouvé dans la classe. Utilisation d'un nom fictif.")

            # Générer un fichier PDF pour le preview
            preview_pdf = generate_pdf(
                links=links,
                selected_class=selected_class,
                subject=subject,
                grade=grade,
                teacher_name=teacher_name,
                student_id=student_id,  # Utiliser l'ID de l'étudiant ou None
                domain=domain,
                assignment_type="Quiz",
                generate_answer_sheet_flag=generate_answer_sheet_flag  # Passer le flag pour inclure la feuille de réponses
            )

            logger.info(f"Fichier PDF généré : {preview_pdf}")

            # Vérifier si le fichier PDF a été généré
            if not os.path.exists(preview_pdf) or os.path.getsize(preview_pdf) == 0:
                raise Exception("Le fichier PDF est vide ou n'a pas été généré.")

            # Retourner le fichier PDF en réponse
            with open(preview_pdf, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'inline; filename="preview.pdf"'
                return response

        except Exception as e:
            logger.error(f"Erreur lors de la génération du preview : {str(e)}")
            return HttpResponse(f"Erreur lors de la génération du preview : {str(e)}", status=500)

    return HttpResponse("Invalid request method.", status=400)

# =============================== #
# Vue Django pour les Elements qui permet les donnEes dynamiques dans la génération   #
# =============================== #

def get_classes(request):
    classes = Class.objects.values('name', 'subject', 'grade')
    return JsonResponse({'classes': list(classes)})

def get_classes_by_grade(request, grade):
    # Filtrer les classes par grade
    classes = Class.objects.filter(grade=grade).values('name', 'subject', 'grade')

    # Retourner la réponse JSON avec les classes filtrées
    return JsonResponse({'classes': list(classes)})

# This is the function that allow us to fetch the subject in the Digital library form
def get_subjects(request):
    """df = pd.read_excel("Digital library.xlsx")"""
    # Nettoyer les espaces dans les sujets
    subjects = ["Math", "Language"]
    return JsonResponse({'subjects': subjects})

# This is the function that allow us to fetch the grade in the Digital library form
def get_grades(request, subject):
    # Sélectionner le bon fichier selon le sujet
    if subject == "Math":
        file_path = "Digital library2.xlsx"
    else:  # Language
        file_path = "Dreametrix excel.xlsx"

    df = pd.read_excel(file_path)

    # Filtrer les grades pour le sujet sélectionné
    filtered_df = df[df["Subject"] == subject]
    grades = filtered_df["Grade"].unique().tolist()

    return JsonResponse({'grades': sorted(grades)})

# API to get Domain based on subject, year, and grade in the digital library form
def get_domains(request, subject, grade):

    # Sélectionner le bon fichier Excel selon le sujet
    if subject == "Math":
        file_path = "Digital library2.xlsx"
    else:  # Language
        file_path = "Dreametrix excel.xlsx"

    df = pd.read_excel(file_path)

    # Filtrer par subject et grade
    filtered_df = df[
        (df["Subject"] == subject) &
        (df["Grade"] == int(grade))
        ]

    # Récupérer les domains uniques

    if filtered_df.empty:
        domains = []  # Handle empty DataFrame case
    else:
        domains = filtered_df["Domain"].unique().tolist()

    return JsonResponse({'domains': domains})

# Specifique Standards
def get_standards(request, subject, grade, domain):
    # Déterminer le bon fichier Excel selon le sujet
    if subject == "Math":
        file_path = "Digital library2.xlsx"
    else:  # Language
        file_path = "Dreametrix excel.xlsx"

    # Lire le fichier Excel
    df = pd.read_excel(file_path)

    # Nettoyer les espaces dans la colonne "Specific Standard"
    df["Specific Standard"] = df["Specific Standard"].str.strip()

    # Filtrer le DataFrame
    filtered_df = df[
        (df["Subject"] == subject) &
        (df["Grade"] == int(grade)) &
        (df["Domain"] == domain)
    ]

    # Récupérer les standards uniques
    if filtered_df.empty:
        available_standards = []  # Gérer le cas où le DataFrame est vide
    else:
        available_standards = filtered_df["Specific Standard"].unique().tolist()

    return JsonResponse({'standards': available_standards})

# API to get_links containings questions based on subject, year, grade and standars in the digital library form
def get_links(request, subject, grade, domain, standards, kind ):
    try:
        # Déterminer le fichier correct en fonction du sujet
        file_path = "Digital library2.xlsx" if subject == "Math" else "Dreametrix excel.xlsx"

        # Charger le fichier Excel
        df = pd.read_excel(file_path)
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Appliquer les filtres
        base_filter = (
            (df["Subject"] == subject) &
            (df["Grade"] == int(grade)) &
            (df["Domain"] == domain) &
            (df["Specific Standard"].isin(standards.split(','))) &  # Utiliser split pour gérer plusieurs standards
            (df["MC/OR"] == kind)
        )
        filtered_df = df.loc[base_filter]

        # Extraire les liens disponibles
        available_links = []
        for index, row in filtered_df.iterrows():
            value = row["Link to item"]
            if pd.notna(value):
                cell = sheet.cell(row=index + 2, column=df.columns.get_loc("Link to item") + 1)
                link = cell.hyperlink.target if cell.hyperlink else None
                if link:
                    available_links.append(link)
        return JsonResponse({'links': available_links})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

#############################################################
#######                                             #########
#######                                             #########
#######                     CLASSES                 #########
#######                                             #########
#######                                             #########
#############################################################

# Calculation Button of the gradebook
def gradebook_calculation(request):
    return render(request, 'dashboard/teacher/calculations.html')

# Function that return the list of all classes in the app
def class_list_view(request):
    classes = Class.objects.all()  # Récupérer toutes les classes

    # Récupérer les filtres de la requête
    class_filter = request.GET.get('class')
    subject_filter = request.GET.get('subject')
    grade_filter = request.GET.get('grade')

    if class_filter:
        classes = classes.filter(name=class_filter)
    if subject_filter:
        classes = classes.filter(subject=subject_filter)
    if grade_filter:
        classes = classes.filter(grade=grade_filter)

    # Récupérer les listes uniques pour les filtres
    unique_classes = Class.objects.values_list('name', flat=True).distinct()
    unique_subjects = Class.objects.values_list('subject', flat=True).distinct()
    unique_grades = Class.objects.values_list('grade', flat=True).distinct()

    return render(request, 'dashboard/teacher/home.html', {
        'classes': classes,
        'unique_classes': unique_classes,
        'unique_subjects': unique_subjects,
        'unique_grades': unique_grades,
    })

# Function to create a new classe
def create_class_view(request):
    """Handle the creation of a new class."""
    if request.method == 'POST':
        name = request.POST.get('name')
        subject = request.POST.get('subject')
        grade = request.POST.get('grade')
        student_ids = request.POST.getlist('students')

        if not name or not subject or not grade:
            return HttpResponseBadRequest("All fields are required.")

        try:
            new_class = Class.objects.create(name=name, subject=subject, grade=grade)
            # Add selected students to the class
            if student_ids:
                valid_student_ids = [int(sid) for sid in student_ids if sid.isdigit()]
                new_class.students.add(*valid_student_ids)

            return redirect('get_classes')
        except Exception as e:
            return HttpResponseBadRequest(f"An error occurred: {e}")
    else:
        # Get all students to display in the form
        students = Student.objects.all()
        return render(request, 'dashboard/teacher/add_new_item_classes.html', {'students': students})

# Function to update a classe
def update_class_view(request, pk):
    """Handle the update of an existing class."""
    class_instance = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        class_instance.name = request.POST.get('name')
        class_instance.subject = request.POST.get('subject')
        class_instance.grade = request.POST.get('grade')
        class_instance.save()
        return redirect('get_classes')  # Redirect to the class list after updating

    return render(request, 'classes/update_class.html', {'class_instance': class_instance})

# Function to delete a classe
def delete_class_view(request, pk):
    """Handle the deletion of a class."""
    class_instance = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        class_instance.delete()
        return redirect('get_classes')  # Redirect to the class list after deletion

    return render(request, 'dashboard/teacher/add_new_item_classes.html', {'class_instance': class_instance})

#############################################################
#######                                             #########
#######                                             #########
#######                   ASSIGNMENT                #########
#######                                             #########
#######                                             #########
#############################################################

from .models import Assignment

def assignment_list_view(request):
    """Render the template to list all assignments."""
    assignments = Assignment.objects.all()
    return render(request, 'assignments/assignment_list.html', {'assignments': assignments})

def create_assignment_view(request):
    """Handle the creation of a new assignment."""
    if request.method == 'POST':
        title = request.POST.get('title')
        assignment_type = request.POST.get('assignment_type')
        average_score = request.POST.get('average_score')
        general_feedback = request.POST.get('general_feedback') == 'on'
        new_assignment = Assignment.objects.create(
            title=title,
            assignment_type=assignment_type,
            average_score=average_score,
            general_feedback=general_feedback
        )
        return redirect('get_assignments')  # Redirect to the assignment list after creation

    return render(request, 'assignments/create_assignment.html')

def update_assignment_view(request, pk):
    """Handle the update of an existing assignment."""
    assignment_instance = get_object_or_404(Assignment, pk=pk)
    if request.method == 'POST':
        assignment_instance.title = request.POST.get('title')
        assignment_instance.assignment_type = request.POST.get('assignment_type')
        assignment_instance.average_score = request.POST.get('average_score')
        assignment_instance.general_feedback = request.POST.get('general_feedback') == 'on'
        assignment_instance.save()
        return redirect('get_assignments')  # Redirect to the assignment list after updating

    return render(request, 'assignments/update_assignment.html', {'assignment_instance': assignment_instance})

def delete_assignment_view(request, pk):
    """Handle the deletion of an assignment."""
    assignment_instance = get_object_or_404(Assignment, pk=pk)
    if request.method == 'POST':
        assignment_instance.delete()
        return redirect('get_assignments')  # Redirect to the assignment list after deletion

    return render(request, 'assignments/delete_assignment.html', {'assignment_instance': assignment_instance})

#############################################################
#######                                             #########
#######                                             #########
#######                  GRADEBOOK                  #########
#######                                             #########
#######                                             #########
#############################################################

from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Case, When, Avg
from django.urls import reverse  # Importation de reverse
from .models import Gradebook, Student, Class

def gradebook_list_menu_classes(request):
    selected_class_id = request.GET.get('class_id')

    if selected_class_id:
        # Rediriger vers la vue gradebook_list_view avec l'ID de la classe sélectionnée
        return redirect(f"{reverse('gradebook_list_view')}?class_id={selected_class_id}")
    else:
        # Mode agrégation globale (non filtré par classe)
        class_data = Gradebook.objects.values(
            'class_instance__name', 'class_instance__id'  # Ajout de l'ID de la classe
        ).annotate(
            total_students=Count('student', distinct=True),
            exam_count=Count(Case(When(assessment_type='EXAM', then=1))),
            test_count=Count(Case(When(assessment_type='TEST', then=1))),
            homework_count=Count(Case(When(assessment_type='HOMEWORK', then=1))),
            class_avg=Avg('score')
        ).order_by('class_instance__name')

        return render(request, 'dashboard/teacher/gradebook_menu.html', {
            'class_data': class_data,
            'classes': Class.objects.all(),
            'has_selected_class': False,
        })

# Function to list all entry of the gradebook
# views.py

def gradebook_list_view(request):
    selected_class_id = request.GET.get('class_id')

    if selected_class_id:
        # Récupération de la classe sélectionnée
        selected_class = get_object_or_404(Class, id=selected_class_id)

        # Récupération des sous-types existants pour chaque type d'évaluation
        exam_subtypes = Gradebook.objects.filter(
            class_instance=selected_class,
            assessment_type='EXAM'
        ).exclude(subtype__isnull=True).values_list('subtype', flat=True).distinct()

        test_subtypes = Gradebook.objects.filter(
            class_instance=selected_class,
            assessment_type='TEST'
        ).exclude(subtype__isnull=True).values_list('subtype', flat=True).distinct()

        homework_subtypes = Gradebook.objects.filter(
            class_instance=selected_class,
            assessment_type='HOMEWORK'
        ).exclude(subtype__isnull=True).values_list('subtype', flat=True).distinct()

        # Récupération des étudiants avec leurs données détaillées
        students = Student.objects.filter(classes=selected_class)
        student_data = []

        for student in students:
            # Calcul de la moyenne générale
            avg = Gradebook.objects.filter(
                student=student,
                class_instance=selected_class
            ).aggregate(average=Avg('score'))['average'] or 0.0

            # Récupération des counts par sous-type
            exam_counts = {
                subtype: Gradebook.objects.filter(
                    student=student,
                    class_instance=selected_class,
                    assessment_type='EXAM',
                    subtype=subtype
                ).count()
                for subtype in exam_subtypes
            }

            test_counts = {
                subtype: Gradebook.objects.filter(
                    student=student,
                    class_instance=selected_class,
                    assessment_type='TEST',
                    subtype=subtype
                ).count()
                for subtype in test_subtypes
            }

            homework_counts = {
                subtype: Gradebook.objects.filter(
                    student=student,
                    class_instance=selected_class,
                    assessment_type='HOMEWORK',
                    subtype=subtype
                ).count()
                for subtype in homework_subtypes
            }

            student_data.append({
                'student': student,
                'average': avg,
                'exam_counts': exam_counts,
                'test_counts': test_counts,
                'homework_counts': homework_counts,
            })

        return render(request, 'dashboard/teacher/gradebook.html', {
            'student_data': student_data,
            'classes': Class.objects.all(),
            'selected_class_id': int(selected_class_id),
            'selected_class': selected_class,  # Ajoutez cette ligne
            'exam_subtypes': list(exam_subtypes),
            'test_subtypes': list(test_subtypes),
            'homework_subtypes': list(homework_subtypes),
            'has_selected_class': True,
        })

    else:
        # Mode agrégation globale (non filtré par classe)
        class_data = Gradebook.objects.values(
            'class_instance__name'
        ).annotate(
            total_students=Count('student', distinct=True),
            exam_count=Count(Case(When(assessment_type='EXAM', then=1))),
            test_count=Count(Case(When(assessment_type='TEST', then=1))),
            homework_count=Count(Case(When(assessment_type='HOMEWORK', then=1))),
            class_avg=Avg('score')
        ).order_by('class_instance__name')

        return render(request, 'dashboard/teacher/gradebook.html', {
            'class_data': class_data,
            'classes': Class.objects.all(),
            'has_selected_class': False,
        })

# Function that return the list of all classes in the app
def get_classes(request):
    classes = Class.objects.values('name', 'subject', 'grade')
    return JsonResponse({'classes': list(classes)})

# views.py
def get_assessment_subtypes(request):
    class_id = request.GET.get('class_id')
    subtypes = Gradebook.objects.filter(
        class_instance_id=class_id
    ).exclude(subtype__isnull=True).values(
        'assessment_type',
        'subtype'
    ).distinct()

    return JsonResponse({'subtypes': list(subtypes)})

# views.py
def get_class_students(request):
    class_id = request.GET.get('class_id')
    students = Student.objects.filter(classes__id=class_id)

    data = []
    for student in students:
        aggregates = Gradebook.objects.filter(
            student=student,
            class_instance_id=class_id
        ).aggregate(
            average=Avg('score'),
            exam_count=Count('id', filter=Q(assessment_type='EXAM')),
            test_count=Count('id', filter=Q(assessment_type='TEST')),
            homework_count=Count('id', filter=Q(assessment_type='HOMEWORK'))
        )

        data.append({
            'id': student.id,
            'name': f"{student.user.first_name} {student.user.last_name}",
            'average': aggregates['average'] or 0.0,
            'exam_count': aggregates['exam_count'],
            'test_count': aggregates['test_count'],
            'homework_count': aggregates['homework_count'],
        })

    return JsonResponse({'students': data})
# Function to update the gradebook entry element
def update_gradebook_view(request, pk):
    """Handle the update of an existing gradebook entry."""
    gradebook_instance = get_object_or_404(Gradebook, pk=pk)
    if request.method == 'POST':
        gradebook_instance.student_name = request.POST.get('student_name')
        gradebook_instance.assignment_title = request.POST.get('assignment_title')
        gradebook_instance.score = request.POST.get('score')
        gradebook_instance.feedback = request.POST.get('feedback')
        gradebook_instance.save()
        return redirect('get_gradebooks')  # Redirect to the gradebook list after updating

    return render(request, 'dashboard/teacher/gradebook.html', {'gradebook_instance': gradebook_instance})

# Function to delete one entry of the grade_book
def delete_gradebook_view(request, pk):
    """Handle the deletion of a gradebook entry."""
    gradebook_instance = get_object_or_404(Gradebook, pk=pk)
    if request.method == 'POST':
        gradebook_instance.delete()
        return redirect('get_gradebooks')  # Redirect to the gradebook list after deletion

    return render(request, 'dashboard/teacher/gradebook.html', {'gradebook_instance': gradebook_instance})

# Function to create the grade book
def create_gradebook_view(request):
    if request.method == 'POST':
        try:
            class_id = request.POST.get('class_instance')
            assessment_type = request.POST.get('assessment_type')
            feedback_file = request.FILES.get('feedback_image')

            if not all([class_id, assessment_type, feedback_file]):
                raise ValueError("Sorry, you need to select all the field")

                Gradebook.objects.create(
                    class_instance_id=class_id,
                    student=student,
                    assessment_type=assessment_type,
                    feedback_file=filename,
                    score=0.0
                )

            return redirect('get_gradebooks')

        except Exception as e:
            classes = Class.objects.all()
            return render(request, 'dashboard/teacher/gradebook_menu.html', {
                'error': str(e),
                'classes': classes,
                'ASSESSMENT_TYPES': Gradebook.ASSESSMENT_TYPES,
            })

    # Code pour les requêtes GET
    classes = Class.objects.all()
    return render(request, 'dashboard/teacher/add_new_item_gradebook.html', {
        'classes': classes,
        'ASSESSMENT_TYPES': Gradebook.ASSESSMENT_TYPES,
    })

# Function to list all student belonging to the selected class
def get_students(request):
    try:
        class_id = request.GET.get('class_id')
        if not class_id:
            return JsonResponse({'error': 'Class ID missing'}, status=400)

        students = Student.objects.filter(classes__id=class_id).values('id', 'user__first_name', 'user__last_name')

        return JsonResponse({
            'students': [
                {
                    'id': s['id'],
                    'name': f"{s['user__first_name']} {s['user__last_name']}"
                }
                for s in students
            ]
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Function to list all sum of the average of each class
def get_average(request):
    class_id = request.GET.get('class_id')
    assessment_type = request.GET.get('assessment_type')

    stats = Gradebook.objects.filter(
        class_instance_id=class_id,
        assessment_type=assessment_type
    ).aggregate(
        average=Avg('score'),
        count=Count('id')
    )

    return JsonResponse({
        'average': stats['average'] if stats['average'] is not None else 0,
        'count': stats['count']
    })


# views.py
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse


@csrf_exempt
def upload_voice_note(request):
    if request.method == 'POST' and request.FILES.get('audio'):
        try:
            student = Student.objects.get(id=request.POST.get('student_id'))
            assessment_type = request.POST.get('assessment_type')
            subtype = request.POST.get('subtype', 'general')

            gradebook_entry, created = Gradebook.objects.update_or_create(
                student=student,
                assessment_type=assessment_type,
                subtype=subtype,
                defaults={'voice_note': request.FILES['audio']}
            )

            return JsonResponse({'status': 'success', 'url': gradebook_entry.voice_note.url})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error'}, status=400)