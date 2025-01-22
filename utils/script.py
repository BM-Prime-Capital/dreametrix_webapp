import pandas as pd
import random
import requests
import fitz
from io import BytesIO
from openpyxl import load_workbook


def filter_question(subject, year, number, grade, kind):
    """
    function to filter the question based on
    subject, year, grade and kind(MC or OR)
    """

    df = pd.read_excel("../Digital library.xlsx")

    filtered_df = df.loc[
        (df["Year"] == year) &
        (df["Subject"] == subject) &
        (df["Grade"] == grade) &
        (df["MC/OR"] == kind)
    ]

    workbook = load_workbook("Digital library.xlsx")
    sheet = workbook.active

    links = []

    column = df.columns.get_loc("Link to item")
    try:
        q = random.sample(list(filtered_df.index), number)
    except ValueError:
        raise ValueError("Not enough questions to generate test")

    for ele in q:
        cell = sheet.cell(ele+2, column+1)
        link = cell.hyperlink.target
        links.append(link)
    return links




def generate_pdf(links: list):
    doc = fitz.open()
    pdf = doc.new_page(width=595, height=842) # size of an A4 sheet (595, 842)
    x, y = 24, 24 # x and y initial coordinate
    margin = 5 # margin


    for link in links:
        link = link.replace("dl=0", "raw=1")

        try:
            response = requests.get(link)
            image = BytesIO(response.content)

            if y+228 > 842-24:
                pdf = doc.new_page(width=595, height=842)
                x, y = 24, 24

            pdf.insert_image(fitz.Rect(x, y, x+228, y+228), stream=image)
            y += 228+margin

        except Exception as e:
            raise ValueError(e)
    doc.save("test.pdf")


#links = filter_question("Math", 2013, 4, grade=5, kind="OR")
#generate_pdf(links)

"""

NB: download the excel file and put it in the same location as this script

 links = filter_question("Math", 2013, 4, grade=5, kind="OR")
 generate_pdf(links)
   
"""
